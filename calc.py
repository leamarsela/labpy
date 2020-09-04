import os
import openpyxl
from openpyxl import load_workbook
from pprint import pprint
from statistics import mean
from math import pi, log10, sqrt


loc = os.getcwd()

wbData = load_workbook('D:\Management\Database\R 00 000 GWL-Database.xlsx')
wsContainer = wbData['Container']
wsRingGamma = wbData['Ring Gamma']
wsGs = wbData['Pycknometer']


def isPycno(id):
    for row in wsGs.values:
        if(row[1] == str(id)):
            weight = row[2]
            valm = row[4]
            valn = row[5]
            return weight, valm, valn
        
def isContainer(id):
    for row in wsContainer.values:
        if(row[1] == id):
            return row[2]

def isRingGamma(id):
    for row in wsRingGamma.values:
        if(row[1] == 1):
            weight = row[2]
            diameter = row[3] / 10.
            thick = row[4] / 10.
            volume = 0.25 * pi * (diameter**2) * thick 
            return weight, diameter, thick, volume

def isNum(num):
    if(num == 'CT'):
        return 'Can not Tested'
    elif(not num):
        return 'Not Tested'
    else:
        try:
            if(float(str(num))):
                return round(float(num),2)
        except ValueError:
            return 'Not Tested'

def reglog(datX, datY, blow):
    log = lambda x:  log10(x)
    square = lambda x: x**2

    logDatX = list(map(log, datX))
    
    meanLogDatX = mean(logDatX)

    sumLogDatX = 0
    for i in logDatX:
        sumLogDatX = sumLogDatX + i

    squareLogDatX = list(map(square, logDatX))

    sumSquareLogDatX = 0
    for i in squareLogDatX:
        sumSquareLogDatX = sumSquareLogDatX + i
    
    meanDatY = mean(datY)

    sumDatY = 0
    for i in datY:
        sumDatY = sumDatY + i
    
    logDatXdatY = []
    for x, y in zip(logDatX, datY):
        logDatXdatY.append(x * y)

    sumLogDatXdatY = 0
    for i in logDatXdatY:
        sumLogDatXdatY = sumLogDatXdatY + i

    n = len(datX)

    valAOne = ((n * sumLogDatXdatY) - (sumLogDatX * sumDatY)) / ((n * sumSquareLogDatX) - (sumLogDatX**2))

    valAZero = meanDatY - (valAOne * meanLogDatX)

    return (log10(blow) * valAOne + valAZero)

def reglinear(datX, datY, valX):

    x = datX[1]
    for i in range(len(datX) - 1):
        if (valX == datX[i]):
            return datY[i]
        elif ((valX - datX[i]) * (valX - x) <= 0 and datX[i] != x):
            a = datY[i - 1] + ((datY[i] - datY[i - 1]) * (valX - x) / (datX[i] - x))
            return a
        x = datX[i]     

def forecast(datX, datY, valX):

    slope = (datY[1] - datY[0]) / (datX[1] - datX[0])

    return (datY[0] - slope * (datX[0] - valX))


# dataWn = []
for file in os.listdir(loc):
    if file.endswith("Classification.xlsx"):
        wb = load_workbook(file)
        ws = wb['Classification']
        
        # calc water content
        tempDataWn = []
        for i in range(1, 4):
            a = {
                'idContWn' + str(i): isNum(ws.cell(1,i).value),
                'wetSoilContWn' + str(i): isNum(ws.cell(2,i).value),
                'drySoilContWn' + str(i): isNum(ws.cell(3,i).value)
                }
            tempDataWn.append(a)
        
        try:
            calcWn = []
            for index in range(3):
                tempCalcWn = ((tempDataWn[index])['wetSoilContWn' + str(index + 1)] - (tempDataWn[index])['drySoilContWn' + str(index + 1)]) * 100 / ((tempDataWn[index])['drySoilContWn' + str(index + 1)] - isContainer((tempDataWn[index])['idContWn' + str(index + 1)]))
                calcWn.append(tempCalcWn)

            wn = round(mean(calcWn), 2)
        except ValueError:
            print('check data water content')


        # calc gamma
        dataGamma = {
            'idRingGamma': isNum(ws.cell(5,1).value),
            'idContGamma': isNum(ws.cell(6,1).value),
            'wetSoilContGamma': isNum(ws.cell(7,1).value),
            'drySoilContGamma': isNum(ws.cell(8,1).value)
            }

        try:
            wetSoil = dataGamma['wetSoilContGamma'] - isContainer(dataGamma['idContGamma']) - isRingGamma((dataGamma['idRingGamma']))[0]
            drySoil = dataGamma['drySoilContGamma'] - isContainer(dataGamma['idContGamma']) - isRingGamma((dataGamma['idRingGamma']))[0]

            wetGamma = wetSoil * 10 / isRingGamma((dataGamma['idRingGamma']))[3]     
            dryGamma = drySoil * 10 / isRingGamma((dataGamma['idRingGamma']))[3]
            wnGamma = (wetSoil - drySoil) * 100 / (drySoil) 
        except ValueError:
            print('check data gamma')


        # calc spesific gravity
        dataGs = []
        for i in range(1, 3):
            a = {
                'idPycno' + str(i): isNum(ws.cell(5,1+i).value),
                'soilPycno' + str(i): isNum(ws.cell(6,1+i).value),
                'waterSoilPycno' + str(i): isNum(ws.cell(7,1+i).value),
                'temperature' + str(i): isNum(ws.cell(8,1+i).value)
                }
            dataGs.append(a)

        try:
            listGs = []
            index = 1
            for i in dataGs:
                tempPycno = isPycno('P' + str(int(i['idPycno' + str(index)])))
                weightPycno = tempPycno[0]
                weightPycnoWater = tempPycno[1] * i['temperature' + str(index)] + tempPycno[2]

                soil = i['soilPycno' + str(index)] - weightPycno

                valGs = soil / (soil + weightPycnoWater - i['waterSoilPycno' + str(index)])

                listGs.append(valGs)

                index+=1
            
            gs = round(mean(listGs), 3)
        except ValueError:
            print('check data spesific gravity')


        #calc Plastic Limit
        tempDataPl = []
        for i in range(1, 4):
            a = {
                'idContPl' + str(i) : isNum(ws.cell(10, i).value),
                'wetSoilContPl' + str(i) : isNum(ws.cell(11, i).value),
                'drySoilContPl' + str(i) : isNum(ws.cell(12, i).value)
            }
            tempDataPl.append(a)
        
        try:
            calcPl = []
            for index in range(3):
                tempCalcPl = ((tempDataPl[index])['wetSoilContPl' + str(index + 1)] - (tempDataPl[index])['drySoilContPl' + str(index + 1)]) * 100 / ((tempDataPl[index])['drySoilContPl' + str(index + 1)] - isContainer((tempDataPl[index])['idContPl' + str(index + 1)]))
                calcPl.append(tempCalcPl)
            
            pl = round(mean(calcPl), 2)
        except ValueError:
            print('check data plastic limit')


        #calc Liquid Limit
        tempDataLl = []
        for i in range(1, 5):
            a = {
                'idContLl' + str(i) : isNum(ws.cell(14, i).value),
                'wetSoilContLl' + str(i) : isNum(ws.cell(15, i).value),
                'numBlow' + str(i) : isNum(ws.cell(16, i).value),
                'drySoilContLl' + str(i) : isNum(ws.cell(17, i).value) 
            }
            tempDataLl.append(a)

        try:
            listWnLl = []
            listNumBlow = []
            for index, data in enumerate(tempDataLl):
                tempWnLl = (data['wetSoilContLl' + str(index + 1)] - data['drySoilContLl' + str(index + 1)]) * 100 / (data['drySoilContLl' + str(index + 1)] - isContainer(data['idContLl' + str(index + 1)]))           
                tempNumBlow = data['numBlow' + str(index + 1)]
                listWnLl.append(tempWnLl)
                listNumBlow.append(tempNumBlow)
    
            ll = reglog(listNumBlow, listWnLl, 25) 
        except ValueError:
            print('check data liquid limit')
        
        
        # calc data sieve
        drySoil = isNum(ws.cell(1,5).value)
        
        try:
            dataSieve = []
            for i in range(1,9):
                a = {'soilSieve' + str(i): isNum(ws.cell(1+i,5).value)}
                dataSieve.append(a)

            sumRetainedSoil = 0
            for index, val in enumerate(dataSieve):
                sumRetainedSoil = sumRetainedSoil + val['soilSieve' + str(index + 1)]

            percentRetained = []
            for i in range(0, 7):
                percentRetained.append(((dataSieve[i])['soilSieve' + str(i + 1)]) * 100 / drySoil)

            pan = (((dataSieve[len(dataSieve) - 1])['soilSieve' + str(len(dataSieve))]) + (drySoil - sumRetainedSoil)) * 100 / drySoil
            percentRetained.append(pan)

            cumulativeRetained = []
            dataCumulative = 0
            for i in percentRetained:
                dataCumulative = dataCumulative + i
                cumulativeRetained.append(dataCumulative)

            percentFiner = []
            for i in cumulativeRetained:
                dataPercentFiner = cumulativeRetained[len(cumulativeRetained) - 1] - i
                percentFiner.append(dataPercentFiner)
        except ValueError:
            print('check data Sieve Analysis')
        

        # calc hydrometer
        status = ws.cell(1, 7).value

        try:
            if(str('Y') == status.upper()):
                weightSample = drySoil - sumRetainedSoil

                constA = (gs * 1.65) / ((gs - 1) * 2.65)

                meniscus = 1.
                zeroCorrection = 0 - meniscus

                time = [1, 2, 4, 8, 15, 30, 45, 90, 180, 1260, 1440]
                
                dataHydro = []
                for i in range(1, 12):
                    if i <= 3:
                        a = {
                            'temp' + str(i): (ws.cell(i,8).value + ws.cell(i,10).value)/2, 
                            'ra' + str(i): (ws.cell(i,9).value + ws.cell(i,11).value)/2,
                            'time' + str(i): time[i - 1]
                            }
                    else:
                        a = {
                            'temp' + str(i): float(ws.cell(i,10).value), 
                            'ra' + str(i): float(ws.cell(i,11).value),
                            'time' + str(i): time[i - 1]
                            }
                    dataHydro.append(a)

                slope = ((dataHydro[1])['ra2'] - (dataHydro[0])['ra1']) / ((dataHydro[1])['time2'] - (dataHydro[0])['time1'])
                valRZero = (dataHydro[0])['ra1'] - (dataHydro[0])['time1'] * slope
                valZero = {
                    'temp0': (dataHydro[0]['temp1']),
                    'ra0': valRZero,
                    'time0': 0
                }
                dataHydro.insert(0, valZero)
                
                listConstT = [15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27, 28, 29, 30]
                listConstCt = [-1.1, -0.9, -0.7, -0.5, -0.3, 0, 0.2, 0.4, 0.7, 1.0, 1.3, 1.65, 2.0, 2.5, 3.05, 3.8]

                listConstR = [i for i in range(61)]
                listConstL = [
                    16.3, 16.1, 16.0, 15.8, 15.6, 15.5, 15.3, 15.2, 15.0, 14.8, 14.7, 
                    14.5, 14.3, 14.2, 14.0, 13.8, 13.7, 13.5, 13.3, 13.2, 13.0,
                    12.9, 12.7, 12.5, 12.4, 12.2, 12.0, 11.9, 11.7, 11.5, 11.4,
                    11.2, 11.1, 10.9, 10.7, 10.6, 10.4, 10.2, 10.1, 9.90, 9.70,
                    9.6, 9.4, 9.2, 9.1, 8.9, 8.8, 8.6, 8.4, 8.3, 8.1,
                    7.9, 7.8, 7.6, 7.4, 7.3, 7.1, 7.0, 6.8, 6.6, 6.5
                    ]
                
                listViscWater = [1.0016e-5, 0.79722e-5]
                listTempViscWater = [20.0, 30.0]

                finerBefore = []
                size = []
                for index, val in enumerate(dataHydro):
                    
                    if (index != 0):
                        valT = val['temp' + str(index)]
                        valCt = reglinear(listConstT, listConstCt, valT)

                        valRct = val['ra' + str(index)] + valCt - zeroCorrection

                        finer = constA * valRct * 100 / weightSample

                        valRcl = val['ra' + str(index)] + meniscus

                        valL = reglinear(listConstR, listConstL, valRcl)

                        valViscosity = reglinear(listTempViscWater, listViscWater, valT)

                        valA = sqrt(30 * valViscosity / (gs - 1))

                        valD = valA * sqrt(valL / val['time' + str(index)])

                        finerBefore.append(finer)

                        size.append(valD)
                        
                dataArrayR = [
                    (dataHydro[1])['ra1'],
                    (dataHydro[2])['ra2'],
                ]
                
                finerO = forecast(dataArrayR, finerBefore[0:2], (dataHydro[0])['ra0'])
                
                finerBefore.insert(0, finerO)
                
                sizeO = forecast(time[0:2], size[0:2], 0)

                size.insert(0, sizeO)

                finerBoundary = forecast(size[0:5], finerBefore[0:5], 0.075)

                correction = percentFiner[-2] / finerBoundary  

                correctionFunc = lambda x: correction * x

                finerAfter = list(map(correctionFunc, finerBefore))
        except ValueError:
            print('check data hydrometer')

